from datetime import datetime

from aiogram.types import InputFile
from openpyxl.workbook import Workbook
from aiogram import types
from aiogram.dispatcher.filters import Text
from pytz import timezone

from data.config import ADMINS
from keyboards.inline.user_inline import confirmation_keyboard_price
from loader import dp, db, bot


@dp.callback_query_handler(Text(startswith='confirmation'), user_id=ADMINS[0])
async def confirmation(call: types.CallbackQuery):
    await call.answer("To'lov tasdiqlangan✅")


@dp.callback_query_handler(Text(startswith='con:'), user_id=ADMINS[0])
async def confirmation(call: types.CallbackQuery):
    user_id = call.data.split(':')[1].strip()
    users_info = await db.select_all_sell_course_users()
    for user_ in users_info:
        if user_[0] == str(user_id):
            if (user_[3] == "Qahramon" or user_[3] == "Бизнес") and user_[4].split(" ")[-1].strip() == \
                    call.message.caption.split(" ")[-5].strip():
                await bot.send_message(chat_id=int(user_id),
                                       text=f"🔰 Qahramon \n\n"
                                            f"Sizning To'lovingiz Tasdiqlandi Opertor siz bn 24soat ichida bog'lanadi",
                                       disable_web_page_preview=True)
                await db.add_con_sell_course_user(
                    id=user_[0],
                    full_name=user_[1],
                    phone_number=user_[2],
                    description=user_[3],
                    course_sell_time=user_[4],
                    user_name=user_[5],
                    price=user_[6],
                    con_time=str(datetime.now(tz=timezone('Asia/Tashkent'))).split('.')[0]
                )
                await call.answer(text="Buyurtma tasdiqlandi✅")
                await call.message.edit_reply_markup(reply_markup=confirmation_keyboard_price)

    for user_ in users_info:
        if user_[0] == str(user_id):
            if (user_[3] == "Qahramon" or user_[3] == "Бизнес") and user_[4].split(" ")[-1].strip() == \
                    call.message.caption.split(" ")[-5].strip():
                await db.add_con_sell_course_user(
                    id=user_[0],
                    full_name=user_[1],
                    phone_number=user_[2],
                    description=user_[3],
                    course_sell_time=user_[4],
                    user_name=user_[5],
                    price=user_[6],
                    con_time=str(datetime.now(tz=timezone('Asia/Tashkent'))).split('.')[0]
                )

    for user in users_info:
        if user[0] == str(user_id):
            if (user[3] == "Chempion" or user[3] == "ВИП") and user[4].split(" ")[-1].strip() == \
                    call.message.caption.split(" ")[-5].strip():
                await bot.send_message(chat_id=int(user_id),
                                       text=f"🔰 Chempion \n\n"
                                            f"Sizning To'lovingiz Tasdiqlandi Opertor siz bn 24soat ichida bog'lanadi",
                                       disable_web_page_preview=True)
                await db.add_con_sell_course_user(
                    id=user[0],
                    full_name=user[1],
                    phone_number=user[2],
                    description=user[3],
                    course_sell_time=user[4],
                    user_name=user[5],
                    price=user[6],
                    con_time=str(datetime.now(tz=timezone('Asia/Tashkent'))).split('.')[0]
                )
                await call.answer(text="Buyurtma tasdiqlandi✅")
                await call.message.edit_reply_markup(reply_markup=confirmation_keyboard_price)


workbook_ = Workbook()
workbook = Workbook()
changed_user_info = workbook.active
not_changed_user = workbook_.active

user_workbook_ = Workbook()
not_user_workbook = Workbook()
all_user_info_ = user_workbook_.active
not_all_user_info = not_user_workbook.active

user_list_workbook = Workbook()
not_user_list_workbook = Workbook()
user_list = user_list_workbook.active
not_user_list = not_user_list_workbook.active


@dp.message_handler(commands=['all_user'], user_id=ADMINS[0])
async def user_get_changed_course(message: types.Message):
    all_users = await db.select_all_users()
    try:
        for z in range(0, len(all_users)):
            changed_user_info[f"A1"] = "Xaridor"
            changed_user_info[f"B1"] = "Foydalanuvchi ID si"
            changed_user_info[f"C1"] = "Foydalanuvchi nomi"
            changed_user_info[f"D1"] = "Ism, Familya"
            changed_user_info[f"E1"] = "Telefon raqami"
            changed_user_info[f"F1"] = "Ro'yxatdan o'tgan vaqti"
            changed_user_info[f"A{z + 2}"] = z + 1
            changed_user_info[f"B{z + 2}"] = all_users[z][0]
            changed_user_info[f"C{z + 2}"] = all_users[z][4]
            changed_user_info[f"D{z + 2}"] = all_users[z][1]
            changed_user_info[f"E{z + 2}"] = all_users[z][2]
            changed_user_info[f"F{z + 2}"] = all_users[z][3]
        workbook.save("user_info.xlsx")
        await bot.send_document(chat_id=message.chat.id, document=InputFile(path_or_bytesio="user_info.xlsx"))
    except:
        not_changed_user[f"A1"] = "Foydalanuvchi topilmadi"
        workbook_.save("not_user_info.xlsx")
        await bot.send_document(chat_id=message.chat.id, document=InputFile(path_or_bytesio="not_user_info.xlsx"))


@dp.message_handler(user_id=ADMINS[0], commands=['course_users'])
async def user_get_all_course(message: types.Message):
    con_course_sell_users = await db.select_all_con_sell_course_users()
    try:
        for z in range(0, len(con_course_sell_users)):
            all_user_info_[f"A1"] = "Xaridor"
            all_user_info_[f"B1"] = "Foydalanuvchi ID si"
            all_user_info_[f"C1"] = "Foydalanuvchi nomi"
            all_user_info_[f"D1"] = "Ism, Familya"
            all_user_info_[f"E1"] = "Telefon raqami"
            all_user_info_[f"F1"] = "Ta'rif"
            all_user_info_[f"G1"] = "To'lov qilingan vaqti"
            all_user_info_[f"H1"] = "To'lov Summasi"
            all_user_info_[f"I1"] = "To'lov tasdiqlangan Vaqt"
            all_user_info_[f"A{z + 2}"] = z + 1
            all_user_info_[f"B{z + 2}"] = con_course_sell_users[z][0]
            all_user_info_[f"C{z + 2}"] = con_course_sell_users[z][5]
            all_user_info_[f"D{z + 2}"] = con_course_sell_users[z][1]
            all_user_info_[f"E{z + 2}"] = con_course_sell_users[z][2]
            all_user_info_[f"F{z + 2}"] = con_course_sell_users[z][3]
            all_user_info_[f"G{z + 2}"] = con_course_sell_users[z][4]
            all_user_info_[f"H{z + 2}"] = con_course_sell_users[z][6]
            all_user_info_[f"I{z + 2}"] = con_course_sell_users[z][7]
        user_workbook_.save("can_sel_course_user_info.xlsx")
        await bot.send_document(chat_id=message.chat.id,
                                document=InputFile(path_or_bytesio="can_sel_course_user_info.xlsx"))
    except:
        not_all_user_info[f"A1"] = "Foydalanuvchi topilmadi"
        not_user_workbook.save("not_can_sel_course_user_info.xlsx")
        await bot.send_document(chat_id=message.chat.id,
                                document=InputFile(path_or_bytesio="not_can_sel_course_user_info.xlsx"))
